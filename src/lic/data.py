"""Deux sources libres : la courbe zéro-coupon de la BdC et la table de mortalité CPM2014 de l'ICA.

La courbe BdC (composition continue, BIS Papers 25) est convertie en taux comptant à
composition ANNUELLE, la convention des taux du chapitre 5 du TSAV : r_annuel = exp(z) - 1.
La table CPM2014 (retraités canadiens, Institut canadien des actuaires) est générationnelle
1999-2013 avec l'échelle CPM-B incorporée ; la colonne 2013, la dernière publiée dans le
classeur, sert de table STATIQUE (précepte déclaré : aucune amélioration au-delà de 2013,
ce qui sous-estime légèrement la duration du passif).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import requests

RAW = Path("data/raw")

CURVE_URL = ("https://www.bankofcanada.ca/stats/results/csv"
             "?lookupPage=lookup_yield_curve.php&startRange=1986-01-01&searchRange=all")
CPM_URL = "https://www.cia-ica.ca/app/themes/wicket/custom/dl_file.php?p=36361&fid=13826"

UA = {"User-Agent": "lic laboratoire pedagogique (github.com/Guilou001/17-licat-alm-ca)"}


def fetch() -> None:
    """Télécharge la courbe et la table (jamais commitées)."""
    RAW.mkdir(parents=True, exist_ok=True)
    for url, name in [(CURVE_URL, "yield_curves.csv"), (CPM_URL, "cpm2014.xlsx")]:
        r = requests.get(url, headers=UA, timeout=300)
        r.raise_for_status()
        (RAW / name).write_bytes(r.content)


def load_curve() -> pd.DataFrame:
    """La courbe quotidienne en composition ANNUELLE (%) : index dates, colonnes années 0,25-30."""
    df = pd.read_csv(RAW / "yield_curves.csv", na_values=[" na", "na"])
    df.columns = [c.strip() for c in df.columns]
    df = df.loc[:, [c for c in df.columns if c == "Date" or c.startswith("ZC")]]
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").astype(float)
    df.columns = [int(c[2:-2]) / 100.0 for c in df.columns]
    return (np.expm1(df) * 100.0).dropna(how="all")        # continue -> annuelle, en %


def load_mortality(sexe: str = "male") -> pd.Series:
    """q_x par âge (18 à 115), colonne 2013 de l'onglet CPM2014 (table statique déclarée)."""
    from openpyxl import load_workbook

    ws = load_workbook(RAW / "cpm2014.xlsx", read_only=True)["CPM2014"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = rows[4]
    years = [(j, c) for j, c in enumerate(header) if isinstance(c, int | float) and c and c > 1900]
    n_par_sexe = len(years) // 2
    bloc = years[:n_par_sexe] if sexe == "male" else years[n_par_sexe:]
    j_2013 = max(bloc, key=lambda t: t[1])[0]
    assert max(bloc, key=lambda t: t[1])[1] == 2013, "la colonne la plus récente doit être 2013"
    qx = {}
    for r in rows[5:]:
        if isinstance(r[0], int | float) and r[0] is not None and r[j_2013] is not None:
            qx[int(r[0])] = float(r[j_2013])
    s = pd.Series(qx).sort_index()
    if not ((s >= 0) & (s <= 1)).all() or s.index[0] != 18 or s.index[-1] != 115:
        raise ValueError("table CPM2014 mal lue")
    return s
